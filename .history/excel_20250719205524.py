import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import boto3
import botocore
from datetime import datetime
import time

def create_cost_sheet_template():
    """สร้าง Excel template สำหรับ Cost Sheet"""
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Failed to create worksheet")
    
    # สร้าง A1:K1 - Cost Sheet
    ws.merge_cells('A1:K1')
    ws['A1'] = "Cost Sheet"
    
    # กำหนด row height = 30
    ws.row_dimensions[1].height = 30
    
    # กำหนด font: Calibri, 26pt, Bold
    font = Font(name="Calibri", size=26, bold=True)
    ws['A1'].font = font
    
    # กำหนด alignment: Left, Middle
    alignment = Alignment(horizontal='left', vertical='center')
    ws['A1'].alignment = alignment
    
    # กำหนดความกว้างของคอลัมน์
    column_widths = {
        'A': 15, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 15,
        'G': 15, 'H': 15, 'I': 15, 'J': 15, 'K': 15
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    return wb

def insert_dynamic_data(wb, data):
    """เพิ่มข้อมูลแบบ dynamic ลงใน template"""
    ws = wb.active
    
    # ตัวอย่างการเพิ่มข้อมูลแบบ dynamic
    # data ควรเป็น dictionary ที่มี key ตามตำแหน่งที่ต้องการใส่ข้อมูล
    # เช่น {"A3": "2024-01-15", "A4": "Project ABC", "D4": "100000"}
    
    for cell_position, value in data.items():
        if cell_position in ws:
            ws[cell_position] = value

def main(columns, rows, minio_access_key, minio_secret_key, bucket_name):
    # ✅ ภายใน Docker network → ใช้ชื่อ container
    minio_internal_endpoint = "http://minio:9000"

    # ✅ สำหรับสร้าง Public Presigned URL → ใช้ Host IP + พอร์ต
    minio_public_endpoint = "http://100.81.135.35:9009"

    now = datetime.now().strftime("%Y%m%d-%H%M%S")
    file_name = f"NCC-{now}.xlsx"
    object_key = f"excel-sheet/{file_name}"

    # 1. สร้าง Excel ใน memory
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Failed to create worksheet")
        
    # สร้าง headers
    for col_num, column_title in enumerate(columns, 1):
        ws[f"{get_column_letter(col_num)}1"] = column_title
    
    # สร้าง data rows
    for row_num, row in enumerate(rows, 2):
        for col_num, column_title in enumerate(columns, 1):
            ws.cell(row=row_num, column=col_num, value=row.get(column_title, ""))
    
    object_key = f"excel-sheet/{file_name}"

    # 2. บันทึกไฟล์ใน memory
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # 3. เชื่อมต่อ MinIO
    config = botocore.config.Config(connect_timeout=30, read_timeout=60)
    s3 = boto3.client(
        "s3",
        endpoint_url=minio_internal_endpoint,  # ✅ ใช้ endpoint ภายใน network ในการอัปโหลดไฟล์
        aws_access_key_id=minio_access_key,
        aws_secret_access_key=minio_secret_key,
        config=config
    )

    # 4. Upload ไฟล์
    excel_buffer.seek(0)
    s3.upload_fileobj(
        excel_buffer,
        bucket_name,
        object_key,
        ExtraArgs={"ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    )

    # 5. สร้าง Presigned URL แบบ public ให้คนอื่นโหลดได้
    public_s3 = boto3.client(
        "s3",
        endpoint_url=minio_public_endpoint,   # ✅ ใช้ Host endpoint เพื่อสร้าง Public URL
        aws_access_key_id=minio_access_key,
        aws_secret_access_key=minio_secret_key,
        config=config
    )

    presigned_url = public_s3.generate_presigned_url(
        'get_object',
        Params={'Bucket': bucket_name, 'Key': object_key},
        ExpiresIn=3600
    )

    return {"file_url": presigned_url}

# ตัวอย่างการใช้งาน
if __name__ == "__main__":
    # ตัวอย่างข้อมูล
    columns = ["Name", "Age", "City", "Salary"]
    rows = [
        {"Name": "John", "Age": 30, "City": "Bangkok", "Salary": 50000},
        {"Name": "Jane", "Age": 25, "City": "Chiang Mai", "Salary": 45000},
        {"Name": "Bob", "Age": 35, "City": "Phuket", "Salary": 60000}
    ]
    
    # เรียกใช้ฟังก์ชัน (ใส่ credentials จริง)
    # result = main(columns, rows, "your_access_key", "your_secret_key", "your_bucket")
    # print(result)
