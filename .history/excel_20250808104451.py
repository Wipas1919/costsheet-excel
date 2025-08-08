"""
EC - AI Cost Estimation System
Excel Generation Module

This module handles the creation of professional cost sheets
for exhibition booth projects using AI-analyzed data.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import json
from datetime import datetime
import os

def create_cost_sheet_template():
    """
    Creates the base Excel template with professional formatting.
    
    Returns:
        openpyxl.Workbook: Configured workbook with template structure
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Header section configuration
    header_data = {
        'A1': 'Cost Sheet',
        'A2': 'Project',
        'E2': 'Booth',
        'K2': 'Decoration',
        'N2': 'Date of cost',
        'A3': 'Showday',
        'E3': 'Budget',
        'N3': 'Designer',
        'A4': 'Place',
        'E4': 'Cost',
        'N4': 'Sales',
        'A5': 'No.',
        'B5': 'Descriptions',
        'E5': 'Size',
        'K5': 'Units',
        'M5': 'Price of unit',
        'N5': 'Amounts',
        'O5': 'Remarks'
    }
    
    # Apply header data
    for cell, value in header_data.items():
        ws[cell] = value
    
    # Apply professional formatting
    apply_template_formatting(ws)
    
    return wb

def apply_template_formatting(ws):
    """
    Applies professional formatting to the worksheet.
    
    Args:
        ws: openpyxl worksheet object
    """
    # Font configurations
    title_font = Font(name='Calibri', size=26, bold=True)
    header_font = Font(name='Calibri', size=11, bold=True)
    
    # Fill configurations
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    total_fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
    
    # Border configuration
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply formatting to headers
    for row in range(1, 6):
        for col in range(1, 16):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.border = thin_border
            if row == 1:
                cell.font = title_font

def process_ai_analyzed_data(columns, rows):
    """
    Processes AI-analyzed data and prepares it for Excel insertion.
    
    Args:
        columns (list): Column headers from AI analysis
        rows (list): Data rows from AI analysis
    
    Returns:
        dict: Processed data ready for Excel generation
    """
    processed_data = {
        'columns': columns,
        'rows': []
    }
    
    for row in rows:
        processed_row = {}
        
        # Map AI data to Excel structure
        for i, col in enumerate(columns):
            if i < len(row):
                processed_row[col] = row[i]
        else:
                processed_row[col] = '-'
        
        # Validate and process dimensions
        processed_row = validate_dimensions(processed_row)
        
        # Calculate quantities based on type
        processed_row = calculate_quantities(processed_row)
        
        processed_data['rows'].append(processed_row)
    
    return processed_data

def validate_dimensions(row_data):
    """
    Validates and processes dimension data.
    
    Args:
        row_data (dict): Row data with W, L, H dimensions
    
    Returns:
        dict: Validated row data
    """
    # Dimension validation logic
    for dim in ['W', 'L', 'H']:
        if dim in row_data:
            try:
                value = float(row_data[dim])
                row_data[dim] = str(value) if value > 0 else '-'
            except (ValueError, TypeError):
                row_data[dim] = '-'
    
    return row_data

def calculate_quantities(row_data):
    """
    Calculates quantities based on component type and dimensions.
    
    Args:
        row_data (dict): Row data with component information
    
    Returns:
        dict: Row data with calculated quantities
    """
    # Quantity calculation logic based on component type
    if 'Component' in row_data and 'Description' in row_data:
        component = row_data['Component']
        description = row_data['Description']
        
        # Calculate based on component type
        if 'Flooring' in component:
            # Flooring calculation logic
            pass
        elif 'Structure' in component:
            # Structure calculation logic
            pass
        elif 'Graphic' in component:
            # Graphic calculation logic
            pass
        elif 'Electrical' in component:
            # Electrical calculation logic
            pass
    
    return row_data

def insert_dynamic_data(ws, processed_data):
    """
    Inserts processed data into the Excel worksheet.
    
    Args:
        ws: openpyxl worksheet object
        processed_data (dict): Processed data from AI analysis
    """
    # Data insertion logic
    start_row = 6  # Starting after headers
    
    for i, row_data in enumerate(processed_data['rows']):
        row_num = start_row + i
        
        # Insert data based on column mapping
        for col_idx, col_name in enumerate(processed_data['columns']):
            cell = ws.cell(row=row_num, column=col_idx + 1)
            cell.value = row_data.get(col_name, '-')
            
            # Apply appropriate formatting
            apply_cell_formatting(cell, col_name, row_data)

def apply_cell_formatting(cell, col_name, row_data):
    """
    Applies appropriate formatting to cells based on content type.
    
    Args:
        cell: openpyxl cell object
        col_name (str): Column name
        row_data (dict): Row data
    """
    # Number formatting for dimensions and prices
    if col_name in ['W', 'L', 'H', 'Quantity', 'price_per_unit', 'total_cost']:
        try:
            value = float(cell.value)
            cell.number_format = '#,##0.00'
        except (ValueError, TypeError):
            pass
    
    # Text wrapping for descriptions
    if col_name in ['Description', 'Component']:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

def generate_cost_sheet(columns, rows):
    """
    Main function to generate cost sheet from AI-analyzed data.
    
    Args:
        columns (list): Column headers from AI analysis
        rows (list): Data rows from AI analysis
    
    Returns:
        dict: Result with file information
    """
    try:
        # Create template
        wb = create_cost_sheet_template()
        ws = wb.active
        
        # Process AI data
        processed_data = process_ai_analyzed_data(columns, rows)
        
        # Insert dynamic data
        insert_dynamic_data(ws, processed_data)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Cost_Sheet_{timestamp}.xlsx'
        
        # Save file (in production, this would upload to cloud storage)
        wb.save(filename)
        
        return {
            'success': True,
            'filename': filename,
            'message': 'Cost sheet generated successfully'
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'message': 'Failed to generate cost sheet'
        }

def main(columns, rows):
    """
    Main entry point for cost sheet generation.
    
    Args:
        columns (list): Column headers
        rows (list): Data rows
    
    Returns:
        dict: Generation result
    """
    return generate_cost_sheet(columns, rows)

# Example usage (for demonstration only)
if __name__ == "__main__":
    # Sample data structure (not actual data)
    sample_columns = [
        'list_id', 'Component', 'Description', 'W', 'L', 'H',
        'Quantity', 'Unit', 'price_per_unit', 'total_cost', 'remark'
    ]
    
    sample_rows = [
        ['100-01-01', 'Flooring', 'Carpet', '10', '5', '0.1', '50', 'sqm', '100', '5000', 'Sample'],
        ['100-02-01', 'Structure', 'Wall Panel', '3', '2', '2.5', '6', 'sqm', '200', '1200', 'Sample']
    ]
    
    result = main(sample_columns, sample_rows)
    print(result)