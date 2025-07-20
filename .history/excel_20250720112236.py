import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import boto3
import botocore
from datetime import datetime
import time

def create_cost_sheet_template():
    """สร้าง Excel template สำหรับ Cost Sheet"""
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Failed to create worksheet")
    
    # A1:K1 - Cost Sheet
    ws.merge_cells('A1:K1')
    ws['A1'] = "Cost Sheet"
    ws.row_dimensions[1].height = 30
    font_a1 = Font(name="Calibri", size=26, bold=True)
    ws['A1'].font = font_a1
    alignment_a1 = Alignment(horizontal='left', vertical='center')
    ws['A1'].alignment = alignment_a1
    
    # A2 - Exhibition Booth
    ws['A2'] = "Exhibition Booth"
    font_a2 = Font(name="Calibri", size=12, color="00AEAAAA")
    ws['A2'].font = font_a2
    alignment_a2 = Alignment(horizontal='left', vertical='center')
    ws['A2'].alignment = alignment_a2
    ws.row_dimensions[2].height = 12
    
    # Font and alignment for template rows
    font10 = Font(name="Calibri", size=10)
    align_left_middle = Alignment(horizontal='left', vertical='center')
    row_height = 12
    
    # A3-A7
    ws['A3'] = "date"
    ws['A4'] = "Project name"
    ws['A5'] = "Place"
    ws['A6'] = "Sales"
    ws['A7'] = "Designer"
    for r in range(3, 8):
        ws[f'A{r}'].font = font10
        ws[f'A{r}'].alignment = align_left_middle
        ws.row_dimensions[r].height = row_height
    
    # D3:F7
    ws.merge_cells('D3:F3')
    ws['D3'] = "Booth"
    ws.merge_cells('D4:F4')
    ws['D4'] = "Budget"
    ws.merge_cells('D5:F5')
    ws['D5'] = "Date of cost"
    ws.merge_cells('D6:F6')
    ws['D6'] = "Estimate By"
    ws.merge_cells('D7:F7')
    ws['D7'] = "Note"
    for r in range(3, 8):
        ws[f'D{r}'].font = font10
        ws[f'D{r}'].alignment = align_left_middle
        ws.row_dimensions[r].height = row_height
    
    # Merge B3:C3, B4:C4, B5:C5, B6:C6, B7:C7
    ws.merge_cells('B3:C3')
    ws.merge_cells('B4:C4')
    ws.merge_cells('B5:C5')
    ws.merge_cells('B6:C6')
    ws.merge_cells('B7:C7')
    for r in range(3, 8):
        ws[f'B{r}'].font = font10
        ws[f'B{r}'].alignment = align_left_middle
    
    # Merge G3:K3, G4:K4, G5:K5, G6:K6, G7:K7
    ws.merge_cells('G3:K3')
    ws.merge_cells('G4:K4')
    ws.merge_cells('G5:K5')
    ws.merge_cells('G6:K6')
    ws.merge_cells('G7:K7')
    for r in range(3, 8):
        ws[f'G{r}'].font = font10
        ws[f'G{r}'].alignment = align_left_middle
    
    # เพิ่มข้อความใน G3 และ G7
    ws['G3'] = "Decoration"
    ws['G7'] = "เป็นราคาที่ใช้ในการประเมิณ"
    
    # Add section headers
    section_headers = [
        (12, "stucture"),
        (15, "Graphic"),
        (18, "Furniture&plant"),
        (21, "Electtrical")
    ]
    font_section = Font(name="Calibri", size=10)
    align_left_middle = Alignment(horizontal='left', vertical='center')
    for row, text in section_headers:
        ws[f'A{row}'] = text
        ws[f'A{row}'].font = font_section
        ws[f'A{row}'].alignment = align_left_middle
    
    # Add B column section items
    b_items = [
        (14, "TOTAL"),
        (17, "TOTAL"),
        (20, "TOTAL"),
        (23, "TOTAL")
    ]
    font_b = Font(name="Calibri", size=10)
    align_left_middle = Alignment(horizontal='left', vertical='center')
    for row, text in b_items:
        ws[f'B{row}'] = text
        ws[f'B{row}'].font = font_b
        ws[f'B{row}'].alignment = align_left_middle
    
    # Set all borders for A3:K7
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(3, 8):
        for col in range(1, 12):  # A=1, K=11
            cell = ws.cell(row=row, column=col)
            cell.border = border
    
    # A8 - Total cost
    ws['A8'] = "Total cost"
    ws.row_dimensions[8].height = 12
    font_a8 = Font(name="Calibri", size=10, color="00FFFFFF")
    fill_a8 = PatternFill(fill_type="solid", start_color="000000", end_color="000000")
    align_a8 = Alignment(horizontal='center', vertical='center')
    ws['A8'].font = font_a8
    ws['A8'].fill = fill_a8
    ws['A8'].alignment = align_a8
    ws.merge_cells('B8:K8')
    fill_white = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
    for col in range(2, 12):  # B=2, K=11
        cell = ws.cell(row=8, column=col)
        cell.font = font_a8
        cell.fill = fill_white
        cell.alignment = align_a8
    # Add border to merged B8:K8
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(2, 12):
        cell = ws.cell(row=8, column=col)
        cell.border = border

    # A11:K11 - Table headers
    headers = [
        ("A11", "Type"),
        ("B11", "Component"),
        ("C11", "Descriptions"),
        ("D11", "W"),
        ("E11", "L"),
        ("F11", "H"),
        ("G11", "quanity"),
        ("H11", "Unit"),
        ("I11", "Unit Price"),
        ("J11", "Amounts"),
        ("K11", "Remarks")
    ]
    font_header = Font(name="Calibri", size=10)
    align_center_middle = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[11].height = 12
    for idx, (cell, text) in enumerate(headers):
        ws[cell] = text
        ws[cell].font = font_header
        ws[cell].alignment = align_center_middle
        # Add border to A11:K11
        col = idx + 1
        ws.cell(row=11, column=col).border = border
    
    # กำหนดความกว้างของคอลัมน์
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 28
    for col in ['D', 'E', 'F']:
        ws.column_dimensions[col].width = 6
    for col in ['G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col].width = 8
    
    return wb

def insert_dynamic_data(wb, data):
    """เพิ่มข้อมูลแบบ dynamic ลงใน template"""
    ws = wb.active
    
    # ตัวอย่างการเพิ่มข้อมูลแบบ dynamic
    # data ควรเป็น dictionary ที่มี key ตามตำแหน่งที่ต้องการใส่ข้อมูล
    # เช่น {"A3": "2024-01-15", "A4": "Project ABC", "D4": "100000"}
    
    for cell_position, value in data.items():
        if cell_position in ws:
            ws[cell_position] = value

def insert_cost_sheet_data(wb, columns, rows):
    """ใส่ข้อมูลจาก columns และ rows ลงใน cost sheet template"""
    ws = wb.active
    
    # ตรวจสอบและใส่ข้อมูลจาก columns
    if rows and len(rows) > 0:
        first_row = rows[0]  # ใช้แถวแรกเป็นตัวแทน
        
        # Project name -> B4
        if "Project name" in columns and "Project name" in first_row:
            ws['B4'] = first_row["Project name"]
        
        # Show day -> B3
        if "Show day" in columns and "Show day" in first_row:
            ws['B3'] = first_row["Show day"]
        
        # Place -> B5
        if "Place" in columns and "Place" in first_row:
            ws['B5'] = first_row["Place"]
    
    # เพิ่มวันที่ปัจจุบันใน G5
    current_date = datetime.now().strftime("%Y-%m-%d")
    ws['G5'] = current_date
    
    # ใส่รายการข้อมูลตาม Type
    insert_item_details(wb, columns, rows)

def insert_item_details(wb, columns, rows):
    """ใส่รายการข้อมูลตาม Type ที่ match กับ template และขยับ template ลง"""
    ws = wb.active
    
    # Column mapping เพื่อให้ตรงกับ template
    column_mapping = {
        'Type': ['Type', 'type', 'TYPE'],
        'Component': ['Component', 'component', 'COMPONENT'],
        'Descriptions': ['Descriptions', 'Description', 'description', 'DESCRIPTION', 'Descriptions'],
        'W': ['W', 'w', 'Width', 'width', 'WIDTH'],
        'L': ['L', 'l', 'Length', 'length', 'LENGTH'],
        'H': ['H', 'h', 'Height', 'height', 'HEIGHT'],
        'quantity': ['quantity', 'Quantity', 'QUANTITY', 'qty', 'Qty', 'QTY'],
        'Unit': ['Unit', 'unit', 'UNIT'],
        'Unit Price': ['Unit Price', 'unit_price', 'Unit_Price', 'price_per_unit', 'Price_per_unit', 'PRICE_PER_UNIT'],
        'Amounts': ['Amounts', 'amounts', 'AMOUNTS', 'total_cost', 'Total_cost', 'TOTAL_COST'],
        'Remarks': ['Remarks', 'remarks', 'REMARKS', 'remark', 'Remark', 'REMARK']
    }
    
    # หา column index ที่ตรงกัน
    column_indices = {}
    for template_col, possible_names in column_mapping.items():
        for col_name in possible_names:
            if col_name in columns:
                column_indices[template_col] = columns.index(col_name)
                break
    
    # Type mapping กับ template
    type_mapping = {
        'structure': 'stucture',
        'Structure': 'stucture',
        'STRUCTURE': 'stucture',
        'graphic': 'Graphic',
        'Graphic': 'Graphic',
        'GRAPHIC': 'Graphic',
        'furniture&plant': 'Furniture&plant',
        'Furniture&plant': 'Furniture&plant',
        'FURNITURE&PLANT': 'Furniture&plant',
        'electtrical': 'Electtrical',
        'Electtrical': 'Electtrical',
        'ELECTTRICAL': 'Electtrical'
    }
    
    # กำหนดตำแหน่งของแต่ละ Type ใน template
    type_rows = {
        'stucture': 12,  # A12
        'Graphic': 15,   # A15
        'Furniture&plant': 18,  # A18
        'Electtrical': 21  # A21
    }
    
    # จัดกลุ่มข้อมูลตาม Type
    type_groups = {}
    for row in rows:
        if 'Type' in column_indices:
            row_type = row[columns[column_indices['Type']]]
            if row_type in type_mapping:
                template_type = type_mapping[row_type]
                if template_type not in type_groups:
                    type_groups[template_type] = []
                type_groups[template_type].append(row)
    
    # เรียงลำดับ Type ตามตำแหน่งใน template (จากล่างขึ้นบน)
    sorted_types = sorted(type_groups.keys(), key=lambda x: type_rows[x], reverse=True)
    
    # ใส่ข้อมูลลงใน template (เริ่มจากล่างขึ้นบนเพื่อไม่ให้กระทบตำแหน่ง)
    for template_type in sorted_types:
        items = type_groups[template_type]
        type_row = type_rows[template_type]
        
        if items:
            # คำนวณจำนวนแถวที่จะเพิ่ม
            num_items = len(items)
            
            # ขยับ template ลงตามจำนวนข้อมูลที่จะเพิ่ม
            # หาแถวสุดท้ายของ template (ก่อน TOTAL)
            if template_type == 'stucture':
                end_row = 14  # ก่อน TOTAL ที่ B14
            elif template_type == 'Graphic':
                end_row = 17  # ก่อน TOTAL ที่ B17
            elif template_type == 'Furniture&plant':
                end_row = 20  # ก่อน TOTAL ที่ B20
            elif template_type == 'Electtrical':
                end_row = 23  # ก่อน TOTAL ที่ B23
            
            # ขยับแถวทั้งหมดจาก end_row ลงไป num_items แถว
            for row in range(end_row, 0, -1):
                for col in range(1, 12):  # A=1, K=11
                    source_cell = ws.cell(row=row, column=col)
                    target_cell = ws.cell(row=row + num_items, column=col)
                    
                    # ตรวจสอบว่า source_cell ไม่ใช่ MergedCell
                    try:
                        # ลองเข้าถึง value เพื่อดูว่าเป็น MergedCell หรือไม่
                        test_value = source_cell.value
                        
                        # คัดลอกค่า
                        target_cell.value = source_cell.value
                        
                        # คัดลอก font (เฉพาะค่าที่จำเป็น)
                        if source_cell.font:
                            target_cell.font = Font(
                                name=source_cell.font.name,
                                size=source_cell.font.size,
                                bold=source_cell.font.bold,
                                italic=source_cell.font.italic,
                                color=source_cell.font.color
                            )
                        
                        # คัดลอก alignment (เฉพาะค่าที่จำเป็น)
                        if source_cell.alignment:
                            target_cell.alignment = Alignment(
                                horizontal=source_cell.alignment.horizontal,
                                vertical=source_cell.alignment.vertical
                            )
                        
                        # คัดลอก border (เฉพาะค่าที่จำเป็น)
                        if source_cell.border:
                            border_sides = {}
                            if source_cell.border.left:
                                border_sides['left'] = source_cell.border.left
                            if source_cell.border.right:
                                border_sides['right'] = source_cell.border.right
                            if source_cell.border.top:
                                border_sides['top'] = source_cell.border.top
                            if source_cell.border.bottom:
                                border_sides['bottom'] = source_cell.border.bottom
                            if border_sides:
                                target_cell.border = Border(**border_sides)
                        
                        # คัดลอก fill (เฉพาะค่าที่จำเป็น)
                        if source_cell.fill and hasattr(source_cell.fill, 'fill_type'):
                            try:
                                target_cell.fill = PatternFill(
                                    fill_type=source_cell.fill.fill_type,
                                    start_color=source_cell.fill.start_color,
                                    end_color=source_cell.fill.end_color
                                )
                            except:
                                # ถ้าไม่สามารถคัดลอก fill ได้ ให้ข้ามไป
                                pass
                        
                        # ล้างค่าเดิม (เฉพาะ cell ที่ไม่ใช่ MergedCell)
                        source_cell.value = None
                        source_cell.font = None
                        source_cell.alignment = None
                        source_cell.border = None
                        source_cell.fill = None
                        
                    except AttributeError:
                        # ถ้าเป็น MergedCell ให้ข้ามไป
                        pass
            
            # อัปเดตตำแหน่ง Type rows ที่เหลือ
            for other_type, other_row in type_rows.items():
                if other_row > type_row:
                    type_rows[other_type] += num_items
            
            # ใส่ข้อมูลใหม่ต่อจาก type row
            for i, item in enumerate(items):
                current_row = type_row + 1 + i
                
                # ใส่ข้อมูลในแต่ละคอลัมน์
                if 'Component' in column_indices:
                    ws[f'B{current_row}'] = item[columns[column_indices['Component']]]
                
                if 'Descriptions' in column_indices:
                    ws[f'C{current_row}'] = item[columns[column_indices['Descriptions']]]
                
                if 'W' in column_indices:
                    ws[f'D{current_row}'] = item[columns[column_indices['W']]]
                
                if 'L' in column_indices:
                    ws[f'E{current_row}'] = item[columns[column_indices['L']]]
                
                if 'H' in column_indices:
                    ws[f'F{current_row}'] = item[columns[column_indices['H']]]
                
                if 'quantity' in column_indices:
                    ws[f'G{current_row}'] = item[columns[column_indices['quantity']]]
                
                if 'Unit' in column_indices:
                    ws[f'H{current_row}'] = item[columns[column_indices['Unit']]]
                
                if 'Unit Price' in column_indices:
                    ws[f'I{current_row}'] = item[columns[column_indices['Unit Price']]]
                
                if 'Amounts' in column_indices:
                    ws[f'J{current_row}'] = item[columns[column_indices['Amounts']]]
                
                if 'Remarks' in column_indices:
                    ws[f'K{current_row}'] = item[columns[column_indices['Remarks']]]
                
                # กำหนด font และ alignment
                font10 = Font(name="Calibri", size=10)
                align_left_middle = Alignment(horizontal='left', vertical='center')
                
                for col in range(2, 12):  # B=2, K=11
                    cell = ws.cell(row=current_row, column=col)
                    cell.font = font10
                    cell.alignment = align_left_middle

def main(columns, rows, minio_access_key, minio_secret_key, bucket_name):
    # ✅ ภายใน Docker network → ใช้ชื่อ container
    minio_internal_endpoint = "http://minio:9000"

    # ✅ สำหรับสร้าง Public Presigned URL → ใช้ Host IP + พอร์ต
    minio_public_endpoint = "http://100.81.135.35:9009"

    now = datetime.now().strftime("%Y%m%d-%H%M%S")
    file_name = f"CostSheet-{now}.xlsx"
    object_key = f"excel-sheet/{file_name}"

    # 1. สร้าง Excel template
    wb = create_cost_sheet_template()
    
    # 2. ใส่ข้อมูลจาก columns และ rows
    insert_cost_sheet_data(wb, columns, rows)

    # 3. บันทึกไฟล์ใน memory
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # 4. เชื่อมต่อ MinIO
    config = botocore.config.Config(connect_timeout=30, read_timeout=60)
    s3 = boto3.client(
        "s3",
        endpoint_url=minio_internal_endpoint,  # ✅ ใช้ endpoint ภายใน network ในการอัปโหลดไฟล์
        aws_access_key_id=minio_access_key,
        aws_secret_access_key=minio_secret_key,
        config=config
    )

    # 5. Upload ไฟล์
    excel_buffer.seek(0)
    s3.upload_fileobj(
        excel_buffer,
        bucket_name,
        object_key,
        ExtraArgs={"ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    )

    # 6. สร้าง Presigned URL แบบ public ให้คนอื่นโหลดได้
    public_s3 = boto3.client(
        "s3",
        endpoint_url=minio_public_endpoint,   # ✅ ใช้ Host endpoint เพื่อสร้าง Public URL
        aws_access_key_id=minio_access_key,
        aws_secret_access_key=minio_secret_key,
        config=config
    )

    presigned_url = public_s3.generate_presigned_url(
        'get_object',
        Params={'Bucket': bucket_name, 'Key': object_key},
        ExpiresIn=3600
    )

    return {"file_url": presigned_url}

# ตัวอย่างการใช้งาน
if __name__ == "__main__":
    # ตัวอย่างข้อมูลสำหรับ Cost Sheet
    columns = ["Project name", "Show day", "Place", "Type", "Component", "Descriptions", "W", "L", "H", "quantity", "Unit", "Unit Price", "Amounts", "Remarks"]
    rows = [
        {
            "Project name": "ABC Exhibition 2024",
            "Show day": "2024-03-15",
            "Place": "Bangkok International Trade & Exhibition Centre",
            "Type": "Structure",
            "Component": "Flooring",
            "Descriptions": "Carpet tiles",
            "W": "3",
            "L": "4",
            "H": "0.02",
            "quantity": "12",
            "Unit": "sqm",
            "Unit Price": "500",
            "Amounts": "6000",
            "Remarks": "Premium quality"
        },
        {
            "Project name": "ABC Exhibition 2024",
            "Show day": "2024-03-15", 
            "Place": "Bangkok International Trade & Exhibition Centre",
            "Type": "Graphic",
            "Component": "Banner",
            "Descriptions": "Vinyl banner 3x4m",
            "W": "3",
            "L": "4",
            "H": "0.1",
            "quantity": "2",
            "Unit": "pcs",
            "Unit Price": "800",
            "Amounts": "1600",
            "Remarks": "Double-sided"
        }
    ]
    
    # เรียกใช้ฟังก์ชัน (ใส่ credentials จริง)
    # result = main(columns, rows, "your_access_key", "your_secret_key", "your_bucket")
    # print(result)
