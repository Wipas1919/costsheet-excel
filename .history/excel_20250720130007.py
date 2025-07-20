import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import boto3
import botocore
from datetime import datetime
import time

def create_cost_sheet_template():
    """สร้าง Excel template สำหรับ Cost Sheet"""
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Failed to create worksheet")
    
    # A1:K1 - Cost Sheet
    ws.merge_cells('A1:K1')
    ws['A1'] = "Cost Sheet"
    ws.row_dimensions[1].height = 30
    font_a1 = Font(name="Calibri", size=26, bold=True)
    ws['A1'].font = font_a1
    alignment_a1 = Alignment(horizontal='left', vertical='center')
    ws['A1'].alignment = alignment_a1
    
    # A2 - Exhibition Booth
    ws['A2'] = "Exhibition Booth"
    font_a2 = Font(name="Calibri", size=12, color="00AEAAAA")
    ws['A2'].font = font_a2
    alignment_a2 = Alignment(horizontal='left', vertical='center')
    ws['A2'].alignment = alignment_a2
    ws.row_dimensions[2].height = 12
    
    # Font and alignment for template rows
    font10 = Font(name="Calibri", size=10)
    align_left_middle = Alignment(horizontal='left', vertical='center')
    row_height = 12
    
    # A3-A7
    ws['A3'] = "date"
    ws['A4'] = "Project name"
    ws['A5'] = "Place"
    ws['A6'] = "Sales"
    ws['A7'] = "Designer"
    for r in range(3, 8):
        ws[f'A{r}'].font = font10
        ws[f'A{r}'].alignment = align_left_middle
        ws.row_dimensions[r].height = row_height
    
    # D3:F7
    ws.merge_cells('D3:F3')
    ws['D3'] = "Booth"
    ws.merge_cells('D4:F4')
    ws['D4'] = "Budget"
    ws.merge_cells('D5:F5')
    ws['D5'] = "Date of cost"
    ws.merge_cells('D6:F6')
    ws['D6'] = "Estimate By"
    ws.merge_cells('D7:F7')
    ws['D7'] = "Note"
    for r in range(3, 8):
        ws[f'D{r}'].font = font10
        ws[f'D{r}'].alignment = align_left_middle
        ws.row_dimensions[r].height = row_height
    
    # Merge B3:C3, B4:C4, B5:C5, B6:C6, B7:C7
    ws.merge_cells('B3:C3')
    ws.merge_cells('B4:C4')
    ws.merge_cells('B5:C5')
    ws.merge_cells('B6:C6')
    ws.merge_cells('B7:C7')
    for r in range(3, 8):
        ws[f'B{r}'].font = font10
        ws[f'B{r}'].alignment = align_left_middle
    
    # Merge G3:K3, G4:K4, G5:K5, G6:K6, G7:K7
    ws.merge_cells('G3:K3')
    ws.merge_cells('G4:K4')
    ws.merge_cells('G5:K5')
    ws.merge_cells('G6:K6')
    ws.merge_cells('G7:K7')
    for r in range(3, 8):
        ws[f'G{r}'].font = font10
        ws[f'G{r}'].alignment = align_left_middle
    
    # Add section headers
    section_headers = []
    font_section = Font(name="Calibri", size=10)
    align_left_middle = Alignment(horizontal='left', vertical='center')
    for row, text in section_headers:
        ws[f'A{row}'] = text
        ws[f'A{row}'].font = font_section
        ws[f'A{row}'].alignment = align_left_middle
    
    # Add B column section items
    b_items = []
    font_b = Font(name="Calibri", size=10)
    align_left_middle = Alignment(horizontal='left', vertical='center')
    for row, text in b_items:
        ws[f'B{row}'] = text
        ws[f'B{row}'].font = font_b
        ws[f'B{row}'].alignment = align_left_middle
    
    # Set all borders for A3:K7
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(3, 8):
        for col in range(1, 12):  # A=1, K=11
            cell = ws.cell(row=row, column=col)
            cell.border = border
    
    # A8 - Total cost
    ws['A8'] = "Total cost"
    ws.row_dimensions[8].height = 12
    font_a8 = Font(name="Calibri", size=10, color="00FFFFFF")
    fill_a8 = PatternFill(fill_type="solid", start_color="000000", end_color="000000")
    align_a8 = Alignment(horizontal='center', vertical='center')
    ws['A8'].font = font_a8
    ws['A8'].fill = fill_a8
    ws['A8'].alignment = align_a8
    ws.merge_cells('B8:K8')
    fill_white = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
    for col in range(2, 12):  # B=2, K=11
        cell = ws.cell(row=8, column=col)
        cell.font = font_a8
        cell.fill = fill_white
        cell.alignment = align_a8
    # Add border to merged B8:K8
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(2, 12):
        cell = ws.cell(row=8, column=col)
        cell.border = border

    # A11:K11 - Table headers
    headers = [
        ("A11", "Type"),
        ("B11", "Component"),
        ("C11", "Descriptions"),
        ("D11", "W"),
        ("E11", "L"),
        ("F11", "H"),
        ("G11", "quanity"),
        ("H11", "Unit"),
        ("I11", "Unit Price"),
        ("J11", "Amounts"),
        ("K11", "Remarks")
    ]
    font_header = Font(name="Calibri", size=10)
    align_center_middle = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[11].height = 12
    for idx, (cell, text) in enumerate(headers):
        ws[cell] = text
        ws[cell].font = font_header
        ws[cell].alignment = align_center_middle
        # Add border to A11:K11
        col = idx + 1
        ws.cell(row=11, column=col).border = border
    
    # กำหนดความกว้างของคอลัมน์
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 28
    for col in ['D', 'E', 'F']:
        ws.column_dimensions[col].width = 6
    for col in ['G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col].width = 8
    
    return wb

def insert_dynamic_data(wb, columns, rows):
    """เพิ่มข้อมูลแบบ dynamic ลงใน template ตาม logic ที่กำหนด"""
    ws = wb.active
    
    # กำหนด type mapping
    type_mapping = {
        '100': 'Structure',
        '101': 'Furniture&plant (For rent & buy out)',
        '102': 'Graphic',
        '103': 'Electrical'
    }
    
    # กำหนด component group mapping
    component_mapping = {
        '01': 'Flooring',
        '02': 'Main structure & decoration'
    }
    
    # ลำดับการแสดงผล type
    type_order = ['100', '101', '102', '103']
    
    # ลำดับการแสดงผล component
    component_order = ['01', '02']
    
    # หา index ของ columns ที่สำคัญ
    list_id_index = columns.index('list_id') if 'list_id' in columns else None
    amounts_index = 9  # column J (index 9)
    
    if list_id_index is None:
        print("Warning: 'list_id' column not found")
        return
    
    # ตรวจสอบว่า rows เป็น list of lists หรือ list of dictionaries
    if not rows:
        print("Warning: No data in rows")
        return
    
    # แปลง rows เป็น list of lists ถ้าเป็น list of dictionaries
    processed_rows = []
    for row in rows:
        if isinstance(row, dict):
            # ถ้าเป็น dictionary ให้แปลงเป็น list ตามลำดับ columns
            row_list = []
            for col in columns:
                row_list.append(row.get(col, ""))
            processed_rows.append(row_list)
        elif isinstance(row, list):
            # ถ้าเป็น list อยู่แล้ว ให้ใช้เลย
            processed_rows.append(row)
        else:
            print(f"Warning: Skipping invalid row format: {type(row)}")
            continue
    
    rows = processed_rows
    
    # เริ่มต้นที่ row 12
    current_row = 12
    
    # จัดกลุ่มข้อมูลตาม type
    data_by_type = {}
    for row_data in rows:
        # ตรวจสอบความยาวของ row_data และ list_id_index
        if not isinstance(row_data, list) or len(row_data) <= list_id_index:
            print(f"Warning: Skipping row with insufficient data: {row_data}")
            continue
            
        list_id = str(row_data[list_id_index])
        if len(list_id) >= 3:
            type_code = list_id[:3]
            if type_code in type_mapping:
                if type_code not in data_by_type:
                    data_by_type[type_code] = []
                data_by_type[type_code].append(row_data)
    
    # วนลูปตามลำดับ type
    for type_code in type_order:
        if type_code not in data_by_type or not data_by_type[type_code]:
            continue
        
        # ใส่ชื่อ type ที่ column A
        ws[f'A{current_row}'] = type_mapping[type_code]
        ws[f'A{current_row}'].font = Font(name="Calibri", size=10)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1
        
        # จัดกลุ่มข้อมูลตาม component
        data_by_component = {}
        for row_data in data_by_type[type_code]:
            # ตรวจสอบความยาวของ row_data และ list_id_index
            if not isinstance(row_data, list) or len(row_data) <= list_id_index:
                continue
                
            list_id = str(row_data[list_id_index])
            if len(list_id) >= 5:
                component_code = list_id[3:5]
                if component_code in component_mapping:
                    if component_code not in data_by_component:
                        data_by_component[component_code] = []
                    data_by_component[component_code].append(row_data)
        
        # วนลูปตามลำดับ component และใส่ใน column B แถวถัดจาก Type
        for component_code in component_order:
            if component_code not in data_by_component or not data_by_component[component_code]:
                continue
            
            # ใส่ชื่อ component group ที่ column B แถวถัดจาก Type
            ws[f'B{current_row}'] = component_mapping[component_code]
            ws[f'B{current_row}'].font = Font(name="Calibri", size=10)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            current_row += 1
        
        # คำนวณ TOTAL สำหรับ type นี้
        total_amount = 0
        for row_data in data_by_type[type_code]:
            # ตรวจสอบความยาวของ row_data และ amounts_index
            if not isinstance(row_data, list) or len(row_data) <= amounts_index:
                continue
                
            try:
                amount = float(row_data[amounts_index]) if row_data[amounts_index] else 0
                total_amount += amount
            except (ValueError, TypeError):
                continue
        
        # ใส่ TOTAL ที่ column B
        ws[f'B{current_row}'] = "TOTAL"
        ws[f'B{current_row}'].font = Font(name="Calibri", size=10)
        ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # ใส่ผลรวมที่ column J
        ws[f'J{current_row}'] = total_amount
        ws[f'J{current_row}'].font = Font(name="Calibri", size=10)
        ws[f'J{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        current_row += 1

def main(columns, rows, minio_access_key, minio_secret_key, bucket_name):
    # ✅ ภายใน Docker network → ใช้ชื่อ container
    minio_internal_endpoint = "http://minio:9000"

    # ✅ สำหรับสร้าง Public Presigned URL → ใช้ Host IP + พอร์ต
    minio_public_endpoint = "http://100.81.135.35:9009"

    now = datetime.now().strftime("%Y%m%d-%H%M%S")
    file_name = f"NCC-{now}.xlsx"
    object_key = f"excel-sheet/{file_name}"

    # 1. สร้าง Excel template
    wb = create_cost_sheet_template()
    
    # 2. เพิ่มข้อมูลแบบ dynamic
    insert_dynamic_data(wb, columns, rows)
    
    object_key = f"excel-sheet/{file_name}"

    # 2. บันทึกไฟล์ใน memory
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # 3. เชื่อมต่อ MinIO
    config = botocore.config.Config(connect_timeout=30, read_timeout=60)
    s3 = boto3.client(
        "s3",
        endpoint_url=minio_internal_endpoint,  # ✅ ใช้ endpoint ภายใน network ในการอัปโหลดไฟล์
        aws_access_key_id=minio_access_key,
        aws_secret_access_key=minio_secret_key,
        config=config
    )

    # 4. Upload ไฟล์
    excel_buffer.seek(0)
    s3.upload_fileobj(
        excel_buffer,
        bucket_name,
        object_key,
        ExtraArgs={"ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    )

    # 5. สร้าง Presigned URL แบบ public ให้คนอื่นโหลดได้
    public_s3 = boto3.client(
        "s3",
        endpoint_url=minio_public_endpoint,   # ✅ ใช้ Host endpoint เพื่อสร้าง Public URL
        aws_access_key_id=minio_access_key,
        aws_secret_access_key=minio_secret_key,
        config=config
    )

    presigned_url = public_s3.generate_presigned_url(
        'get_object',
        Params={'Bucket': bucket_name, 'Key': object_key},
        ExpiresIn=3600
    )

    return {"file_url": presigned_url}

# ตัวอย่างการใช้งาน
if __name__ == "__main__":
    # ตัวอย่างข้อมูลสำหรับ Cost Sheet
    columns = ["list_id", "Project name", "Show day", "Place", "type", "Component", "Description", "W", "L", "H", "Unit", "Quantity", "price_per_unit", "total_cost", "remark"]
    rows = [
        ["10001", "Project A", "2024-01-15", "Bangkok", "Structure", "Flooring", "Carpet", "10", "5", "0.1", "m2", "50", "100", "5000", "High quality"],
        ["10002", "Project A", "2024-01-15", "Bangkok", "Structure", "Main structure", "Wall", "3", "2", "2.5", "m2", "6", "200", "1200", "Standard"],
        ["10101", "Project A", "2024-01-15", "Bangkok", "Furniture", "Chair", "Office chair", "0.5", "0.5", "1", "pcs", "10", "500", "5000", "Ergonomic"],
        ["10201", "Project A", "2024-01-15", "Bangkok", "Graphic", "Banner", "Vinyl banner", "2", "1", "0.1", "m2", "2", "150", "300", "UV resistant"]
    ]
    
    # เรียกใช้ฟังก์ชัน (ใส่ credentials จริง)
    # result = main(columns, rows, "your_access_key", "your_secret_key", "your_bucket")
    # print(result)
