import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import boto3
import botocore
from datetime import datetime
import time

def create_cost_sheet_template():
    """Create Excel template for Cost Sheet"""
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Failed to create worksheet")
    
    # A1:K1 - Cost Sheet
    ws.merge_cells('A1:K1')
    ws['A1'] = "Cost Sheet"
    ws.row_dimensions[1].height = 30
    font_a1 = Font(name="Calibri", size=26, bold=True)
    ws['A1'].font = font_a1
    alignment_a1 = Alignment(horizontal='left', vertical='center')
    ws['A1'].alignment = alignment_a1
    
    # A2 - Exhibition Booth
    ws['A2'] = "Exhibition Booth"
    font_a2 = Font(name="Calibri", size=12, color="00AEAAAA")
    ws['A2'].font = font_a2
    alignment_a2 = Alignment(horizontal='left', vertical='center')
    ws['A2'].alignment = alignment_a2
    ws.row_dimensions[2].height = 12
    
    # Font and alignment for template rows
    font10 = Font(name="Calibri", size=10)
    align_left_middle = Alignment(horizontal='left', vertical='center')
    row_height = 12
    
    # A3-A7
    ws['A3'] = "date"
    ws['A4'] = "Project name"
    ws['A5'] = "Place"
    ws['A6'] = "Sales"
    ws['A7'] = "Designer"
    for r in range(3, 8):
        ws[f'A{r}'].font = font10
        ws[f'A{r}'].alignment = align_left_middle
        ws.row_dimensions[r].height = row_height
    
    # D3:F7
    ws.merge_cells('D3:F3')
    ws['D3'] = "Booth"
    ws.merge_cells('D4:F4')
    ws['D4'] = "Budget"
    ws.merge_cells('D5:F5')
    ws['D5'] = "Date of cost"
    ws.merge_cells('D6:F6')
    ws['D6'] = "Estimate By"
    ws.merge_cells('D7:F7')
    ws['D7'] = "Note"
    for r in range(3, 8):
        ws[f'D{r}'].font = font10
        ws[f'D{r}'].alignment = align_left_middle
        ws.row_dimensions[r].height = row_height
    
    # Merge B3:C3, B4:C4, B5:C5, B6:C6, B7:C7
    ws.merge_cells('B3:C3')
    ws.merge_cells('B4:C4')
    ws.merge_cells('B5:C5')
    ws.merge_cells('B6:C6')
    ws.merge_cells('B7:C7')
    for r in range(3, 8):
        ws[f'B{r}'].font = font10
        ws[f'B{r}'].alignment = align_left_middle
    
    # Merge G3:K3, G4:K4, G5:K5, G6:K6, G7:K7
    ws.merge_cells('G3:K3')
    ws.merge_cells('G4:K4')
    ws.merge_cells('G5:K5')
    ws.merge_cells('G6:K6')
    ws.merge_cells('G7:K7')
    
    # Set text in G3, G5 and G7
    ws['G3'] = "Decoration"
    
    # Set current date in G5
    current_date = datetime.now().strftime("%d/%m/%Y")
    ws['G5'] = current_date
    
    ws['G7'] = "This price is an estimate"
    
    for r in range(3, 8):
        ws[f'G{r}'].font = font10
        ws[f'G{r}'].alignment = align_left_middle
    
    # Add section headers
    section_headers = []
    font_section = Font(name="Calibri", size=10)
    align_left_middle = Alignment(horizontal='left', vertical='center')
    for row, text in section_headers:
        ws[f'A{row}'] = text
        ws[f'A{row}'].font = font_section
        ws[f'A{row}'].alignment = align_left_middle
    
    # Add B column section items
    b_items = []
    font_b = Font(name="Calibri", size=10)
    align_left_middle = Alignment(horizontal='left', vertical='center')
    for row, text in b_items:
        ws[f'B{row}'] = text
        ws[f'B{row}'].font = font_b
        ws[f'B{row}'].alignment = align_left_middle
    
    # Set all borders for A3:K7
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(3, 8):
        for col in range(1, 12):  # A=1, K=11
            cell = ws.cell(row=row, column=col)
            cell.border = border
    
    # A8 - Total cost
    ws['A8'] = "Total cost"
    ws.row_dimensions[8].height = 12
    font_a8 = Font(name="Calibri", size=10, color="00FFFFFF")
    fill_a8 = PatternFill(fill_type="solid", start_color="000000", end_color="000000")
    align_a8 = Alignment(horizontal='center', vertical='center')
    ws['A8'].font = font_a8
    ws['A8'].fill = fill_a8
    ws['A8'].alignment = align_a8
    ws.merge_cells('B8:K8')
    fill_white = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
    for col in range(2, 12):  # B=2, K=11
        cell = ws.cell(row=8, column=col)
        cell.font = font_a8
        cell.fill = fill_white
        cell.alignment = align_a8
    # Add border to merged B8:K8
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(2, 12):
        cell = ws.cell(row=8, column=col)
        cell.border = border
    
    # Add All Borders to A8
    ws['A8'].border = border

    # A11:K11 - Table headers
    headers = [
        ("A11", "Type"),
        ("B11", "Component"),
        ("C11", "Descriptions"),
        ("D11", "W"),
        ("E11", "L"),
        ("F11", "H"),
        ("G11", "quantity"),
        ("H11", "Unit"),
        ("I11", "Unit Price"),
        ("J11", "Amounts"),
        ("K11", "Remarks")
    ]
    font_header = Font(name="Calibri", size=10)
    align_center_middle = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[11].height = 12
    for idx, (cell, text) in enumerate(headers):
        ws[cell] = text
        ws[cell].font = font_header
        ws[cell].alignment = align_center_middle
        # Add border to A11:K11
        col = idx + 1
        ws.cell(row=11, column=col).border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 28
    for col in ['D', 'E', 'F']:
        ws.column_dimensions[col].width = 6
    for col in ['G', 'H']:
        ws.column_dimensions[col].width = 8
    # Expand columns I, J to width = 9
    ws.column_dimensions['I'].width = 9
    ws.column_dimensions['J'].width = 9
    ws.column_dimensions['K'].width = 8
    
    return wb

def insert_dynamic_data(wb, columns, rows):
    """Insert dynamic data into template according to specified logic"""
    ws = wb.active
    
    # Define type mapping
    type_mapping = {
        '100': 'Structure',
        '101': 'Furniture&plant (For rent & buy out)',
        '102': 'Graphic',
        '103': 'Electrical'
    }
    
    # Define component group mapping
    component_mapping = {
        '01': 'Flooring',
        '02': 'Main structure & decoration'
    }
    
    # Type display order
    type_order = ['100', '101', '102', '103']
    
    # Component display order
    component_order = ['01', '02']
    
    # Find important column indices
    list_id_index = columns.index('list_id') if 'list_id' in columns else None
    amounts_index = 9  # column J (index 9)
    
    if list_id_index is None:
        print("Warning: 'list_id' column not found")
        return
    
    # Check if rows is list of lists or list of dictionaries
    if not rows:
        print("Warning: No data in rows")
        return
    
    # Convert rows to list of lists if it's list of dictionaries
    processed_rows = []
    for row in rows:
        if isinstance(row, dict):
            # If it's a dictionary, convert to list according to column order
            row_list = []
            for col in columns:
                row_list.append(row.get(col, ""))
            processed_rows.append(row_list)
        elif isinstance(row, list):
            # If it's already a list, use it directly
            processed_rows.append(row)
        else:
            print(f"Warning: Skipping invalid row format: {type(row)}")
            continue
    
    rows = processed_rows
    
    # Insert data from columns into template
    if rows:
        # Insert Project name into B4
        if 'Project name' in columns:
            project_name_idx = columns.index('Project name')
            for row_data in rows:
                if isinstance(row_data, list) and len(row_data) > project_name_idx and row_data[project_name_idx]:
                    ws['B4'] = row_data[project_name_idx]
                    break
        
        # Insert Show day into B3
        if 'Show day' in columns:
            show_day_idx = columns.index('Show day')
            for row_data in rows:
                if isinstance(row_data, list) and len(row_data) > show_day_idx and row_data[show_day_idx]:
                    ws['B3'] = row_data[show_day_idx]
                    break
        
        # Insert Place into B5
        if 'Place' in columns:
            place_idx = columns.index('Place')
            for row_data in rows:
                if isinstance(row_data, list) and len(row_data) > place_idx and row_data[place_idx]:
                    ws['B5'] = row_data[place_idx]
                    break
    
    # Start at row 12
    current_row = 12
    
    # Store data for calculating grand total
    grand_total = 0
    type_start_rows = {}  # Store starting row of each type
    total_rows = []  # Store TOTAL rows
    
    # Group data by type
    data_by_type = {}
    for row_data in rows:
        # Check row_data length and list_id_index
        if not isinstance(row_data, list) or len(row_data) <= list_id_index:
            print(f"Warning: Skipping row with insufficient data: {row_data}")
            continue
            
        list_id = str(row_data[list_id_index])
        print(f"Debug: list_id = {list_id}")
        if len(list_id) >= 3:
            type_code = list_id[:3]
            print(f"Debug: type_code = {type_code}")
            if type_code in type_mapping:
                if type_code not in data_by_type:
                    data_by_type[type_code] = []
                data_by_type[type_code].append(row_data)
            else:
                print(f"Debug: type_code {type_code} not found in mapping")
    
    # Loop through types in order
    for type_code in type_order:
        if type_code not in data_by_type or not data_by_type[type_code]:
            continue
        
        # Insert type name in column A
        type_name = type_mapping[type_code]
        ws[f'A{current_row}'] = type_name
        ws[f'A{current_row}'].font = Font(name="Calibri", size=10)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # If exceeds 15 characters, wrap to new line
        if len(type_name) > 15:
            # Find suitable break point (space, parentheses, ampersand, hyphen)
            break_point = 15
            for i in range(15, 0, -1):
                if type_name[i] in [' ', '(', '&', '-']:
                    break_point = i + 1
                    break
            
            # Create text with line break
            wrapped_text = type_name[:break_point] + '\n' + type_name[break_point:]
            ws[f'A{current_row}'] = wrapped_text
        type_start_rows[type_code] = current_row  # Store starting row of this type
        current_row += 1
        
        # Group data by component
        data_by_component = {}
        for row_data in data_by_type[type_code]:
            # Check row_data length and list_id_index
            if not isinstance(row_data, list) or len(row_data) <= list_id_index:
                continue
                
            list_id = str(row_data[list_id_index])
            print(f"Debug: Processing component for list_id = {list_id}")
            if len(list_id) >= 5:
                component_code = list_id[3:5]
                print(f"Debug: component_code = {component_code}")
                if component_code in component_mapping:
                    if component_code not in data_by_component:
                        data_by_component[component_code] = []
                    data_by_component[component_code].append(row_data)
                else:
                    print(f"Debug: component_code {component_code} not found in mapping")
        
        # Loop through components in order and insert in column B next row after Type
        for component_code in component_order:
            if component_code not in data_by_component or not data_by_component[component_code]:
                continue
            
            # Insert component group name in column B next row after Type
            ws[f'B{current_row}'] = component_mapping[component_code]
            ws[f'B{current_row}'].font = Font(name="Calibri", size=10)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            current_row += 1
        
        # If no matching component found, display all data
        if not any(component_code in data_by_component for component_code in component_order):
            # Insert all data in this type
            for row_data in data_by_type[type_code]:
                if not isinstance(row_data, list):
                    continue
                
                # Insert data in next row
                if 'Component' in columns:
                    comp_idx = columns.index('Component')
                    if comp_idx < len(row_data):
                        ws[f'B{current_row}'] = row_data[comp_idx]
                        ws[f'B{current_row}'].font = Font(name="Calibri", size=10)
                        ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
                
                # Insert Description in column C
                if 'Description' in columns:
                    desc_idx = columns.index('Description')
                    if desc_idx < len(row_data):
                        ws[f'C{current_row}'] = row_data[desc_idx]
                        ws[f'C{current_row}'].font = Font(name="Calibri", size=10)
                        ws[f'C{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
                
                # Insert W in column D
                if 'W' in columns:
                    w_idx = columns.index('W')
                    if w_idx < len(row_data):
                        ws[f'D{current_row}'] = row_data[w_idx]
                        ws[f'D{current_row}'].font = Font(name="Calibri", size=10)
                        ws[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Insert L in column E
                if 'L' in columns:
                    l_idx = columns.index('L')
                    if l_idx < len(row_data):
                        ws[f'E{current_row}'] = row_data[l_idx]
                        ws[f'E{current_row}'].font = Font(name="Calibri", size=10)
                        ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Insert H in column F
                if 'H' in columns:
                    h_idx = columns.index('H')
                    if h_idx < len(row_data):
                        ws[f'F{current_row}'] = row_data[h_idx]
                        ws[f'F{current_row}'].font = Font(name="Calibri", size=10)
                        ws[f'F{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Insert Quantity in column G
                if 'Quantity' in columns:
                    qty_idx = columns.index('Quantity')
                    if qty_idx < len(row_data):
                        ws[f'G{current_row}'] = row_data[qty_idx]
                        ws[f'G{current_row}'].font = Font(name="Calibri", size=10)
                        ws[f'G{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Insert Unit in column H
                if 'Unit' in columns:
                    unit_idx = columns.index('Unit')
                    if unit_idx < len(row_data):
                        ws[f'H{current_row}'] = row_data[unit_idx]
                        ws[f'H{current_row}'].font = Font(name="Calibri", size=10)
                        ws[f'H{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Insert price_per_unit in column I
                if 'price_per_unit' in columns:
                    price_idx = columns.index('price_per_unit')
                    if price_idx < len(row_data):
                        ws[f'I{current_row}'] = row_data[price_idx]
                        ws[f'I{current_row}'].font = Font(name="Calibri", size=10)
                        ws[f'I{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                        # Add comma style for Unit Price
                        try:
                            if row_data[price_idx]:
                                ws[f'I{current_row}'].number_format = '#,##0.00'
                        except:
                            pass
                
                # Insert total_cost in column J
                if 'total_cost' in columns:
                    total_idx = columns.index('total_cost')
                    if total_idx < len(row_data):
                        ws[f'J{current_row}'] = row_data[total_idx]
                        ws[f'J{current_row}'].font = Font(name="Calibri", size=10)
                        ws[f'J{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                        # Add comma style for Amounts
                        try:
                            if row_data[total_idx]:
                                ws[f'J{current_row}'].number_format = '#,##0.00'
                        except:
                            pass
                
                # Insert remark in column K
                if 'remark' in columns:
                    remark_idx = columns.index('remark')
                    if remark_idx < len(row_data):
                        ws[f'K{current_row}'] = row_data[remark_idx]
                        ws[f'K{current_row}'].font = Font(name="Calibri", size=10)
                        ws[f'K{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                current_row += 1
        
        # Calculate TOTAL for this type
        total_amount = 0
        print(f"Debug: Calculating total for type {type_code}")
        for row_data in data_by_type[type_code]:
            # Check row_data length
            if not isinstance(row_data, list):
                continue
                
            # Find total_cost index
            if 'total_cost' in columns:
                total_cost_idx = columns.index('total_cost')
                if total_cost_idx < len(row_data):
                    try:
                        amount = float(row_data[total_cost_idx]) if row_data[total_cost_idx] else 0
                        print(f"Debug: Adding amount {amount} from total_cost")
                        total_amount += amount
                    except (ValueError, TypeError):
                        print(f"Debug: Error converting {row_data[total_cost_idx]} to float")
                        continue
        print(f"Debug: Total amount for type {type_code} = {total_amount}")
        
        # Add this type's total to grand total
        grand_total += total_amount
        
        # Insert TOTAL in column B
        ws[f'B{current_row}'] = "TOTAL"
        ws[f'B{current_row}'].font = Font(name="Calibri", size=10)
        ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Insert total in column J
        ws[f'J{current_row}'] = total_amount
        ws[f'J{current_row}'].font = Font(name="Calibri", size=10)
        ws[f'J{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        # Add comma style for TOTAL Amounts
        ws[f'J{current_row}'].number_format = '#,##0.00'
        
        # Add background color #808080 to columns B:J in TOTAL row
        gray_fill = PatternFill(fill_type="solid", start_color="808080", end_color="808080")
        for col in range(2, 11):  # B=2, J=10
            cell = ws.cell(row=current_row, column=col)
            cell.fill = gray_fill
        
        total_rows.append(current_row)  # Store TOTAL row
        current_row += 1
    
    # Insert grand total in B8
    ws['B8'] = grand_total
    ws['B8'].font = Font(name="Calibri", size=10, color="000000")  # black color
    ws['B8'].alignment = Alignment(horizontal='center', vertical='center')
    # Add comma style for Grand Total
    ws['B8'].number_format = '#,##0.00'
    
    # Add borders as in the image - vertical lines by column and horizontal lines at end of total
    thin = Side(border_style="thin", color="000000")
    
    for type_code in type_order:
        if type_code in type_start_rows:
            start_row = type_start_rows[type_code] - 1  # row with type name
            
            # Find correct end_row for this type
            end_row = current_row - 1  # default to last row
            for check_type in type_order:
                if check_type == type_code:
                    break
                if check_type in type_start_rows:
                    end_row = type_start_rows[check_type] - 2  # row before next type
            
            # Add borders as in image
            for row in range(start_row, end_row + 1):
                # Top row of type - top and vertical lines
                if row == start_row:
                    for col in range(1, 12):  # A to K
                        cell = ws.cell(row=row, column=col)
                        if col == 1:  # Column A
                            cell.border = Border(top=thin, left=thin, right=thin)
                        elif col == 11:  # Column K
                            cell.border = Border(top=thin, right=thin)
                        else:  # Column B-J
                            cell.border = Border(top=thin, right=thin)
                
                # Bottom row of type (TOTAL) - bottom and vertical lines
                elif row == end_row:
                    for col in range(1, 12):  # A to K
                        cell = ws.cell(row=row, column=col)
                        if col == 1:  # Column A
                            cell.border = Border(left=thin, right=thin, bottom=thin)
                        elif col == 11:  # Column K
                            cell.border = Border(right=thin, bottom=thin)
                        else:  # Column B-J
                            cell.border = Border(right=thin, bottom=thin)
                
                # Middle rows - only vertical lines
                else:
                    for col in range(1, 12):  # A to K
                        cell = ws.cell(row=row, column=col)
                        if col == 1:  # Column A
                            cell.border = Border(left=thin, right=thin)
                        elif col == 11:  # Column K
                            cell.border = Border(right=thin)
                        else:  # Column B-J
                            cell.border = Border(right=thin)
    
    # Add horizontal lines under all TOTAL rows
    for total_row in total_rows:
        for col in range(1, 12):  # A to K
            cell = ws.cell(row=total_row, column=col)
            current_border = cell.border
            cell.border = Border(
                top=current_border.top,
                left=current_border.left,
                right=current_border.right,
                bottom=thin
            )
    
    # Add All borders to A11:K11 (table headers) - add at the end to ensure visibility
    for col in range(1, 12):  # A to K
        cell = ws.cell(row=11, column=col)
        # Add All borders (top, left, right, bottom)
        cell.border = Border(
            top=thin,
            left=thin,
            right=thin,
            bottom=thin
        )

def main(columns, rows, minio_access_key, minio_secret_key, bucket_name):
    # Configure your MinIO endpoints here
    minio_internal_endpoint = "YOUR_INTERNAL_MINIO_ENDPOINT"
    minio_public_endpoint = "YOUR_PUBLIC_MINIO_ENDPOINT"

    now = datetime.now().strftime("%Y%m%d-%H%M%S")
    file_name = f"CostSheet-{now}.xlsx"
    object_key = f"excel-sheet/{file_name}"

    # 1. Create Excel template
    wb = create_cost_sheet_template()
    
    # 2. Add dynamic data
    insert_dynamic_data(wb, columns, rows)
    
    object_key = f"excel-sheet/{file_name}"

    # 2. Save file in memory
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # 3. Connect to MinIO
    config = botocore.config.Config(connect_timeout=30, read_timeout=60)
    s3 = boto3.client(
        "s3",
        endpoint_url=minio_internal_endpoint,
        aws_access_key_id=minio_access_key,
        aws_secret_access_key=minio_secret_key,
        config=config
    )

    # 4. Upload file
    excel_buffer.seek(0)
    s3.upload_fileobj(
        excel_buffer,
        bucket_name,
        object_key,
        ExtraArgs={"ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    )

    # 5. Create public Presigned URL for others to download
    public_s3 = boto3.client(
        "s3",
        endpoint_url=minio_public_endpoint,
        aws_access_key_id=minio_access_key,
        aws_secret_access_key=minio_secret_key,
        config=config
    )

    presigned_url = public_s3.generate_presigned_url(
        'get_object',
        Params={'Bucket': bucket_name, 'Key': object_key},
        ExpiresIn=3600
    )

    return {"file_url": presigned_url}

# Usage example
if __name__ == "__main__":
    # Sample data for demonstration
    columns = ["list_id", "Project name", "Show day", "Place", "type", "Component", "Description", "W", "L", "H", "Unit", "Quantity", "price_per_unit", "total_cost", "remark"]
    rows = [
        ["10001", "Sample Project", "2024-01-15", "Sample Location", "Structure", "Flooring", "Sample Material", "10", "5", "0.1", "m2", "50", "100", "5000", "Sample remark"],
        ["10002", "Sample Project", "2024-01-15", "Sample Location", "Structure", "Main structure", "Sample Wall", "3", "2", "2.5", "m2", "6", "200", "1200", "Sample remark"],
        ["10101", "Sample Project", "2024-01-15", "Sample Location", "Furniture", "Chair", "Sample Chair", "0.5", "0.5", "1", "pcs", "10", "500", "5000", "Sample remark"],
        ["10201", "Sample Project", "2024-01-15", "Sample Location", "Graphic", "Banner", "Sample Banner", "2", "1", "0.1", "m2", "2", "150", "300", "Sample remark"]
    ]
    
    # Call function with your credentials
    # result = main(columns, rows, "your_access_key", "your_secret_key", "your_bucket")
    # print(result)