import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import boto3
import botocore
from datetime import datetime
import time

def create_cost_sheet_template():
    """สร้าง Excel template สำหรับ Cost Sheet"""
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Failed to create worksheet")
    
    # กำหนด template data
    template_data = [
        # Cell, Value, MergedRange, FontName, FontSize, Bold, Italic, Underline, FontColorHex, BackgroundHex, BorderLeft, BorderTop, BorderRight, BorderBottom
        ("A1", "Cost Sheet", "A1:K1", "Calibri", 26.0, True, False, False, "", "", "none", "none", "none", "none"),
        ("A2", "Exhibition Booth", "A2:C2", "Calibri", 12.0, False, False, False, "", "", "none", "none", "none", "none"),
        ("A3", "date", "", "Calibri", 10.0, False, False, False, "", "", "thin", "thin", "thin", "thin"),
        ("B3", "", "B3:C3", "Calibri", 10.0, False, False, False, "", "", "thin", "thin", "thin", "thin"),
        ("D3", "Booth", "D3:F3", "Calibri", 10.0, False, False, False, "", "", "thin", "thin", "thin", "thin"),
        ("G3", "", "G3:K3", "Calibri", 10.0, False, False, False, "", "", "thin", "thin", "thin", "thin"),
        ("A4", "Project name", "", "Calibri", 10.0, False, False, False, "", "", "thin", "thin", "thin", "thin"),
        ("B4", "", "B4:C4", "Calibri", 10.0, False, False, False, "", "", "thin", "thin", "thin", "thin"),
        ("D4", "Budget", "D4:F4", "Calibri", 10.0, False, False, False, "", "", "thin", "thin", "thin", "thin"),
        ("G4", "", "G4:K4", "Calibri", 10.0, False, False, False, "", "", "thin", "thin", "thin", "thin"),
    ]
    
    # เพิ่ม rows ว่างสำหรับข้อมูล (A5-K20)
    for row in range(5, 21):
        template_data.extend([
            (f"A{row}", "", f"A{row}:C{row}", "Calibri", 10.0, False, False, False, "", "", "thin", "thin", "thin", "thin"),
            (f"D{row}", "", f"D{row}:F{row}", "Calibri", 10.0, False, False, False, "", "", "thin", "thin", "thin", "thin"),
            (f"G{row}", "", f"G{row}:K{row}", "Calibri", 10.0, False, False, False, "", "", "thin", "thin", "thin", "thin"),
        ])
    
    # สร้าง template
    for cell, value, merged_range, font_name, font_size, bold, italic, underline, font_color, bg_color, border_left, border_top, border_right, border_bottom in template_data:
        # กำหนดค่าใน cell
        ws[cell] = value
        
        # กำหนด font
        font = Font(
            name=font_name,
            size=font_size,
            bold=bold,
            italic=italic,
            underline='single' if underline else None,
            color=font_color if font_color else None
        )
        ws[cell].font = font
        
        # กำหนด background color
        if bg_color:
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            ws[cell].fill = fill
        
        # กำหนด borders
        border_sides = {}
        if border_left != "none":
            border_sides['left'] = Side(style='thin' if border_left == 'thin' else 'medium')
        if border_top != "none":
            border_sides['top'] = Side(style='thin' if border_top == 'thin' else 'medium')
        if border_right != "none":
            border_sides['right'] = Side(style='thin' if border_right == 'thin' else 'medium')
        if border_bottom != "none":
            border_sides['bottom'] = Side(style='thin' if border_bottom == 'thin' else 'medium')
        
        if border_sides:
            border = Border(**border_sides)
            ws[cell].border = border
        
        # จัดการ alignment สำหรับ merged cells
        if merged_range:
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        
        # จัดการ merged cells (ทำหลังจากกำหนด style แล้ว)
        if merged_range:
            ws.merge_cells(merged_range)
    
    # กำหนดความกว้างของคอลัมน์
    column_widths = {
        'A': 15, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 15,
        'G': 15, 'H': 15, 'I': 15, 'J': 15, 'K': 15
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    return wb

def insert_dynamic_data(wb, data):
    """เพิ่มข้อมูลแบบ dynamic ลงใน template"""
    ws = wb.active
    
    # ตัวอย่างการเพิ่มข้อมูลแบบ dynamic
    # data ควรเป็น dictionary ที่มี key ตามตำแหน่งที่ต้องการใส่ข้อมูล
    # เช่น {"A3": "2024-01-15", "A4": "Project ABC", "D4": "100000"}
    
    for cell_position, value in data.items():
        if cell_position in ws:
            ws[cell_position] = value

def main(columns, rows, minio_access_key, minio_secret_key, bucket_name, template_type="cost_sheet", dynamic_data=None):
    # ✅ ภายใน Docker network → ใช้ชื่อ container
    minio_internal_endpoint = "http://minio:9000"

    # ✅ สำหรับสร้าง Public Presigned URL → ใช้ Host IP + พอร์ต
    minio_public_endpoint = "http://100.81.135.35:9009"

    now = datetime.now().strftime("%Y%m%d-%H%M%S")
    
    # 1. สร้าง Excel ตาม template type
    if template_type == "cost_sheet":
        file_name = f"CostSheet-{now}.xlsx"
        wb = create_cost_sheet_template()
        
        # เพิ่มข้อมูลแบบ dynamic (ถ้ามี)
        if dynamic_data:
            insert_dynamic_data(wb, dynamic_data)
    else:
        # สร้าง Excel แบบเดิม (dynamic columns และ rows)
        file_name = f"NCC-{now}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise ValueError("Failed to create worksheet")
            
        # สร้าง headers
        for col_num, column_title in enumerate(columns, 1):
            ws[f"{get_column_letter(col_num)}1"] = column_title
        
        # สร้าง data rows
        for row_num, row in enumerate(rows, 2):
            for col_num, column_title in enumerate(columns, 1):
                ws.cell(row=row_num, column=col_num, value=row.get(column_title, ""))
    
    object_key = f"excel-sheet/{file_name}"

    # 3. บันทึกไฟล์ใน memory
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # 4. เชื่อมต่อ MinIO
    config = botocore.config.Config(connect_timeout=30, read_timeout=60)
    s3 = boto3.client(
        "s3",
        endpoint_url=minio_internal_endpoint,  # ✅ ใช้ endpoint ภายใน network ในการอัปโหลดไฟล์
        aws_access_key_id=minio_access_key,
        aws_secret_access_key=minio_secret_key,
        config=config
    )

    # 5. Upload ไฟล์
    excel_buffer.seek(0)
    s3.upload_fileobj(
        excel_buffer,
        bucket_name,
        object_key,
        ExtraArgs={"ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    )

    # 6. สร้าง Presigned URL แบบ public ให้คนอื่นโหลดได้
    public_s3 = boto3.client(
        "s3",
        endpoint_url=minio_public_endpoint,   # ✅ ใช้ Host endpoint เพื่อสร้าง Public URL
        aws_access_key_id=minio_access_key,
        aws_secret_access_key=minio_secret_key,
        config=config
    )

    presigned_url = public_s3.generate_presigned_url(
        'get_object',
        Params={'Bucket': bucket_name, 'Key': object_key},
        ExpiresIn=3600
    )

    return {"file_url": presigned_url}

# ตัวอย่างการใช้งาน
if __name__ == "__main__":
    # ตัวอย่างข้อมูลแบบ dynamic
    sample_data = {
        "A3": "2024-01-15",
        "A4": "Project ABC Exhibition",
        "D4": "500,000",
        "A5": "Booth Design",
        "D5": "100,000",
        "A6": "Construction",
        "D6": "200,000",
        "A7": "Lighting",
        "D7": "50,000",
        "A8": "Total",
        "D8": "350,000"
    }
    
    # เรียกใช้ฟังก์ชัน (ใส่ credentials จริง)
    # result = main("your_access_key", "your_secret_key", "your_bucket", sample_data)
    # print(result)
