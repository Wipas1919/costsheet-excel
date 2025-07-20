import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import boto3
import botocore
from datetime import datetime
import time

def create_cost_sheet_template():
    """สร้าง Excel template สำหรับ Cost Sheet"""
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Failed to create worksheet")
    
    # A1:K1 - Cost Sheet
    ws.merge_cells('A1:K1')
    ws['A1'] = "Cost Sheet"
    ws.row_dimensions[1].height = 30
    font_a1 = Font(name="Calibri", size=26, bold=True)
    ws['A1'].font = font_a1
    alignment_a1 = Alignment(horizontal='left', vertical='center')
    ws['A1'].alignment = alignment_a1
    
    # A2 - Exhibition Booth
    ws['A2'] = "Exhibition Booth"
    font_a2 = Font(name="Calibri", size=12, color="00AEAAAA")
    ws['A2'].font = font_a2
    alignment_a2 = Alignment(horizontal='left', vertical='center')
    ws['A2'].alignment = alignment_a2
    ws.row_dimensions[2].height = 12
    
    # Font and alignment for template rows
    font10 = Font(name="Calibri", size=10)
    align_left_middle = Alignment(horizontal='left', vertical='center')
    row_height = 12
    
    # A3-A7
    ws['A3'] = "date"
    ws['A4'] = "Project name"
    ws['A5'] = "Place"
    ws['A6'] = "Sales"
    ws['A7'] = "Designer"
    for r in range(3, 8):
        ws[f'A{r}'].font = font10
        ws[f'A{r}'].alignment = align_left_middle
        ws.row_dimensions[r].height = row_height
    
    # D3:F7
    ws.merge_cells('D3:F3')
    ws['D3'] = "Booth"
    ws.merge_cells('D4:F4')
    ws['D4'] = "Budget"
    ws.merge_cells('D5:F5')
    ws['D5'] = "Date of cost"
    ws.merge_cells('D6:F6')
    ws['D6'] = "Estimate By"
    ws.merge_cells('D7:F7')
    ws['D7'] = "Note"
    for r in range(3, 8):
        ws[f'D{r}'].font = font10
        ws[f'D{r}'].alignment = align_left_middle
        ws.row_dimensions[r].height = row_height
    
    # Merge B3:C3, B4:C4, B5:C5, B6:C6, B7:C7
    ws.merge_cells('B3:C3')
    ws.merge_cells('B4:C4')
    ws.merge_cells('B5:C5')
    ws.merge_cells('B6:C6')
    ws.merge_cells('B7:C7')
    for r in range(3, 8):
        ws[f'B{r}'].font = font10
        ws[f'B{r}'].alignment = align_left_middle
    
    # Merge G3:K3, G4:K4, G5:K5, G6:K6, G7:K7
    ws.merge_cells('G3:K3')
    ws.merge_cells('G4:K4')
    ws.merge_cells('G5:K5')
    ws.merge_cells('G6:K6')
    ws.merge_cells('G7:K7')
    for r in range(3, 8):
        ws[f'G{r}'].font = font10
        ws[f'G{r}'].alignment = align_left_middle
    
    # Add section headers
    section_headers = [
        (12, "stucture"),
        (16, "Graphic"),
        (20, "Furniture&plant"),
        (22, "Electtrical")
    ]
    font_section = Font(name="Calibri", size=10)
    align_left_middle = Alignment(horizontal='left', vertical='center')
    for row, text in section_headers:
        ws[f'A{row}'] = text
        ws[f'A{row}'].font = font_section
        ws[f'A{row}'].alignment = align_left_middle
    
    # Add B column section items
    b_items = [
        (13, "Flooring"),
        (14, "Main struce & decoration"),
        (15, "TOTAL"),
        (17, "Main struce & decoration"),
        (18, "TOTAL"),
        (20, "Main struce & decoration"),
        (21, "TOTAL"),
        (23, "Main struce & decoration"),
        (24, "TOTAL")
    ]
    font_b = Font(name="Calibri", size=10)
    align_left_middle = Alignment(horizontal='left', vertical='center')
    for row, text in b_items:
        ws[f'B{row}'] = text
        ws[f'B{row}'].font = font_b
        ws[f'B{row}'].alignment = align_left_middle
    
    # Set all borders for A3:K7
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(3, 8):
        for col in range(1, 12):  # A=1, K=11
            cell = ws.cell(row=row, column=col)
            cell.border = border
    
    # A8 - Total cost
    ws['A8'] = "Total cost"
    ws.row_dimensions[8].height = 12
    font_a8 = Font(name="Calibri", size=10, color="00FFFFFF")
    fill_a8 = PatternFill(fill_type="solid", start_color="000000", end_color="000000")
    align_a8 = Alignment(horizontal='center', vertical='center')
    ws['A8'].font = font_a8
    ws['A8'].fill = fill_a8
    ws['A8'].alignment = align_a8
    ws.merge_cells('B8:K8')
    fill_white = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
    for col in range(2, 12):  # B=2, K=11
        cell = ws.cell(row=8, column=col)
        cell.font = font_a8
        cell.fill = fill_white
        cell.alignment = align_a8
    # Add border to merged B8:K8
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(2, 12):
        cell = ws.cell(row=8, column=col)
        cell.border = border

    # A11:K11 - Table headers
    headers = [
        ("A11", "Type"),
        ("B11", "Component"),
        ("C11", "Descriptions"),
        ("D11", "W"),
        ("E11", "L"),
        ("F11", "H"),
        ("G11", "quanity"),
        ("H11", "Unit"),
        ("I11", "Unit Price"),
        ("J11", "Amounts"),
        ("K11", "Remarks")
    ]
    font_header = Font(name="Calibri", size=10)
    align_center_middle = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[11].height = 12
    for idx, (cell, text) in enumerate(headers):
        ws[cell] = text
        ws[cell].font = font_header
        ws[cell].alignment = align_center_middle
        # Add border to A11:K11
        col = idx + 1
        ws.cell(row=11, column=col).border = border
    
    # กำหนดความกว้างของคอลัมน์
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 28
    for col in ['D', 'E', 'F']:
        ws.column_dimensions[col].width = 6
    for col in ['G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col].width = 8
    
    return wb

def insert_dynamic_data(wb, data):
    """เพิ่มข้อมูลแบบ dynamic ลงใน template"""
    ws = wb.active
    
    # ตัวอย่างการเพิ่มข้อมูลแบบ dynamic
    # data ควรเป็น dictionary ที่มี key ตามตำแหน่งที่ต้องการใส่ข้อมูล
    # เช่น {"A3": "2024-01-15", "A4": "Project ABC", "D4": "100000"}
    
    for cell_position, value in data.items():
        if cell_position in ws:
            ws[cell_position] = value

def main(columns, rows, minio_access_key, minio_secret_key, bucket_name):
    # ✅ ภายใน Docker network → ใช้ชื่อ container
    minio_internal_endpoint = "http://minio:9000"

    # ✅ สำหรับสร้าง Public Presigned URL → ใช้ Host IP + พอร์ต
    minio_public_endpoint = "http://100.81.135.35:9009"

    now = datetime.now().strftime("%Y%m%d-%H%M%S")
    file_name = f"NCC-{now}.xlsx"
    object_key = f"excel-sheet/{file_name}"

    # 1. สร้าง Excel template
    wb = create_cost_sheet_template()
    
    object_key = f"excel-sheet/{file_name}"

    # 2. บันทึกไฟล์ใน memory
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # 3. เชื่อมต่อ MinIO
    config = botocore.config.Config(connect_timeout=30, read_timeout=60)
    s3 = boto3.client(
        "s3",
        endpoint_url=minio_internal_endpoint,  # ✅ ใช้ endpoint ภายใน network ในการอัปโหลดไฟล์
        aws_access_key_id=minio_access_key,
        aws_secret_access_key=minio_secret_key,
        config=config
    )

    # 4. Upload ไฟล์
    excel_buffer.seek(0)
    s3.upload_fileobj(
        excel_buffer,
        bucket_name,
        object_key,
        ExtraArgs={"ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    )

    # 5. สร้าง Presigned URL แบบ public ให้คนอื่นโหลดได้
    public_s3 = boto3.client(
        "s3",
        endpoint_url=minio_public_endpoint,   # ✅ ใช้ Host endpoint เพื่อสร้าง Public URL
        aws_access_key_id=minio_access_key,
        aws_secret_access_key=minio_secret_key,
        config=config
    )

    presigned_url = public_s3.generate_presigned_url(
        'get_object',
        Params={'Bucket': bucket_name, 'Key': object_key},
        ExpiresIn=3600
    )

    return {"file_url": presigned_url}

# ตัวอย่างการใช้งาน
if __name__ == "__main__":
    # ตัวอย่างข้อมูล (ยังไม่ได้ใช้ใน template ปัจจุบัน)
    columns = ["Name", "Age", "City", "Salary"]
    rows = [
        {"Name": "John", "Age": 30, "City": "Bangkok", "Salary": 50000},
        {"Name": "Jane", "Age": 25, "City": "Chiang Mai", "Salary": 45000},
        {"Name": "Bob", "Age": 35, "City": "Phuket", "Salary": 60000}
    ]
    
    # เรียกใช้ฟังก์ชัน (ใส่ credentials จริง)
    # result = main(columns, rows, "your_access_key", "your_secret_key", "your_bucket")
    # print(result)
